import selenium
import time
import datetime
import glob
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
from openpyxl import load_workbook



driver = webdriver.Chrome()
wait = WebDriverWait(driver, 60)
driver.get("https://influence.sshrc-crsh.gc.ca/200002/1001/Lists/FAR2020/UserView.aspx")
Username = driver.find_element_by_id("ctl00_PlaceHolderMain_signInControl_UserName")
Password = driver.find_element_by_id("ctl00_PlaceHolderMain_signInControl_Password")
Sign_In = driver.find_element_by_id("ctl00_PlaceHolderMain_signInControl_LoginButton")
Username.send_keys("LOGIN-EMAIL")
Password.send_keys("LOGIN-PASSWORD")
Sign_In.click()

action = ActionChains(driver)

wb = load_workbook(r'LOCATION OF EXCEL SPREADSHEET') 
ws = wb.active

i = 1
n = 0

row_num = ws.max_row + 1

#for i in range(2,24):
for i in range(2, row_num):
    n = n + 1
    Agency = ws.cell(row = i, column = 2)
    Last_Name = ws.cell(row = i, column = 3)
    First_Name = ws.cell(row = i, column = 4)
    Department = ws.cell(row = i, column = 5)
    Start_Date = ws.cell(row = i, column = 6)
    if Last_Name.value is None:
        break
    
    driver.get('https://influence.sshrc-crsh.gc.ca/200002/1001/Lists/FAR2020/UserView.aspx?View={953A89BE-8CC3-4714-924F-FA82D155EF21}&FilterField1=Agency&FilterValue1=' + Agency.value + '&FilterField2=Last%5Fx0020%5Fname%5Fx0020%5Fof%5Fx0020%5FA&FilterValue2=' + Last_Name.value + '&FilterField3=First%5Fx0020%5FName%5Fx0020%5Fof%5Fx0020%5F&FilterValue3=' + First_Name.value)
    Edit_Button = driver.find_element_by_xpath('//*[@alt="Edit"]')
    Edit_Button.click()
    Award_Status_Field = driver.find_element_by_xpath('//*[@title="Award status Required Field"]')
    Department_Field = driver.find_element_by_xpath('//*[@title="Department Required Field"]')
    Reg_Status_Field = driver.find_element_by_xpath('//*[@title="Registration status Required Field"]')
    Elig_Criteria_Field = driver.find_element_by_xpath('//*[@title="All eligibility criteria met? Required Field"]')
    Start_Date_Field = driver.find_element_by_xpath('//*[@title="Start date of award Required Field"]')
    End_Date_Field = driver.find_element_by_xpath('//*[@title="End date of award Required Field"]')

    Award_Status_Field.send_keys('Accepted')
    
    Department_Field.click()
    Department_Field.send_keys(Keys.CONTROL + "a")
    Department_Field.send_keys(Keys.DELETE)
    Department_Field.send_keys(Department.value)
    Reg_Status_Field.send_keys('Full-Time')
    Elig_Criteria_Field.send_keys('Yes')
    Start_Date_Field.click()
    Start_Date_Field.send_keys(Keys.CONTROL + "a")
    Start_Date_Field.send_keys(Keys.DELETE)
    End_Date_Field.click()
    End_Date_Field.send_keys(Keys.CONTROL + "a")
    End_Date_Field.send_keys(Keys.DELETE)



    if Start_Date.value == 'May 2020':
        Start_Date_Field.send_keys('01/05/2020')
        End_Date_Field.send_keys('4/30/2021')
    if Start_Date.value == 'September 2020':
        Start_Date_Field.send_keys('01/09/2020')
        End_Date_Field.send_keys('8/31/2021')
    if Start_Date.value == 'January 2021':
        Start_Date_Field.send_keys('01/01/2021')
        End_Date_Field.send_keys('12/31/2021')

    Save_Button = End_Date_Field = driver.find_element_by_xpath('//*[@value="Save"]')
    Save_Button.click()


